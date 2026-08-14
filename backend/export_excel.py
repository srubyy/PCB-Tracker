import sys
import json
import re
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def extract_mfg_year(barcode, explicit_year=None):
    if explicit_year is not None and str(explicit_year).strip():
        try:
            yr = int(str(explicit_year).strip())
            if 2000 <= yr <= 2050:
                return yr
        except ValueError:
            pass

    if not barcode or barcode == '-' or len(barcode) < 4:
        return None
    s = str(barcode).strip()
    if s.startswith('AT') and len(s) <= 8:
        return None
    if s.lower().startswith('sa') and len(s) <= 8:
        return None

    # Direct 4-digit year match (2010 to 2050)
    four_digit = re.search(r'(20[1-5]\d)', s)
    if four_digit:
        yr = int(four_digit.group(1))
        if 2010 <= yr <= 2050:
            return yr
    
    # Regex matching: any letter followed by exactly 2 digits (e.g. B22, E26, D21)
    matches = re.findall(r'[a-zA-Z](\d{2})', s)
    if matches:
        for m in matches:
            yr = int(m)
            if 10 <= yr <= 50:
                return 2000 + yr
                
    # Fallback: standard 3rd/4th character check
    if len(s) >= 4:
        yr_part = s[2:4]
        try:
            yr = int(yr_part)
            if 10 <= yr <= 50:
                return 2000 + yr
        except ValueError:
            pass
            
    # Legacy fallbacks
    if len(s) in [16, 17]:
        try:
            yr = int(s[3:5])
            return yr + 2000
        except ValueError:
            pass
    if s.startswith('AGV'):
        c_idx = s.find('C')
        if c_idx != -1 and c_idx + 2 < len(s):
            try:
                yr = int(s[c_idx+1 : c_idx+3])
                return yr + 2000
            except ValueError:
                pass
    if s.startswith('EA') and len(s) == 22:
        try:
            yr = int(s[15:17])
            return yr + 2000
        except ValueError:
            pass
    return None

def find_column_indices(rows):
    dummy_idx = -1
    barcode_idx = -1
    mfg_year_idx = -1
    
    if len(rows) > 0:
        header = [str(x).lower().strip() for x in rows[0]]
        for idx, col in enumerate(header):
            if any(term in col for term in ['pcb sr no', 'pcb serial', 'dummy', 'sr no', 'sr_no']):
                if dummy_idx == -1:
                    dummy_idx = idx
            if any(term in col for term in ['barcode', 'actual serial', 'real serial']):
                if barcode_idx == -1:
                    barcode_idx = idx
            if any(term in col for term in ['mfg year', 'year', 'manufacturing year']):
                if mfg_year_idx == -1:
                    mfg_year_idx = idx
    return dummy_idx, barcode_idx, mfg_year_idx

def generate_excel(json_data):
    dest_path = json_data['dest_file_path']
    raw_sheets = json_data['raw_sheets']
    cell_edits = json_data['cell_edits']
    scan_logs = json_data.get('scan_logs', [])
    export_history = json_data.get('export_history', [])
    
    scrap_year_threshold = json_data.get('scrap_year_threshold')
    separate_year_threshold = json_data.get('separate_year_threshold')
    checkbox_year_threshold = json_data.get('checkbox_year_threshold')
    
    # Initialize ExcelWriter
    with pd.ExcelWriter(dest_path, engine='openpyxl') as writer:
        # Rule 1: Preserve all original sheets exactly, with edits overlaid
        for sheet_name, rows in raw_sheets.items():
            if not rows:
                continue
                
            # Filter edits for this sheet
            sheet_edits = [e for e in cell_edits if e['sheet_name'] == sheet_name]
            
            # Find column indices
            dummy_col_idx, barcode_col_idx, mfg_year_col_idx = find_column_indices(rows)
            
            # Apply edits where col_idx is numeric index
            for r_idx in range(len(rows)):
                for edit in sheet_edits:
                    if int(edit['row_idx']) == r_idx:
                        try:
                            c_idx = int(edit['col_idx'])
                            if 0 <= c_idx < len(rows[r_idx]):
                                rows[r_idx][c_idx] = edit['value']
                        except ValueError:
                            pass # Virtual edits handled separately below
            
            # Process rows to insert virtual columns to the immediate right of dummyColIdx
            new_rows = []
            
            # Header row modification
            header_row = list(rows[0])
            insert_pos = dummy_col_idx + 1 if dummy_col_idx != -1 else 1
            
            virtual_headers = [
                "Actual Serial No",
                "Length of Actual Serial No",
                "Mfg Year",
                "Action"
            ]
            
            for i, v_h in enumerate(virtual_headers):
                header_row.insert(insert_pos + i, v_h)
            new_rows.append(header_row)
            
            # Data rows modification
            for r_idx in range(1, len(rows)):
                row = list(rows[r_idx])
                
                # Fetch actual serial barcode
                actual_barcode = ''
                # Check for virtual edit first
                for edit in sheet_edits:
                    if int(edit['row_idx']) == r_idx and String_Match(edit['col_idx'], 'actual_serial_no'):
                        actual_barcode = edit['value']
                # Fallback to barcode column if no edit
                if not actual_barcode and barcode_col_idx != -1 and barcode_col_idx < len(row):
                    actual_barcode = row[barcode_col_idx]
                
                actual_barcode = str(actual_barcode).strip()
                if actual_barcode == '-':
                    actual_barcode = ''
                    
                # Compute virtual column values
                barcode_length = len(actual_barcode) if actual_barcode else 0
                raw_mfg_year_val = row[mfg_year_col_idx] if (mfg_year_col_idx != -1 and mfg_year_col_idx < len(row)) else None
                calculated_year = extract_mfg_year(actual_barcode, raw_mfg_year_val)
                
                repairable_val = 'No'
                for edit in sheet_edits:
                    if int(edit['row_idx']) == r_idx and String_Match(edit['col_idx'], 'repairable'):
                        repairable_val = 'Yes' if edit['value'] == 'true' else 'No'
                
                action_val = 'Pending' if not actual_barcode else '-'
                if calculated_year:
                    scrap_limit = scrap_year_threshold if scrap_year_threshold is not None else 2021
                    sep_limit = separate_year_threshold if separate_year_threshold is not None else 2022
                    chk_limit = checkbox_year_threshold if checkbox_year_threshold is not None else 2023
                    
                    if calculated_year <= scrap_limit:
                        action_val = 'Scrap'
                    elif separate_year_threshold is not None and calculated_year == sep_limit:
                        action_val = 'Separate'
                    elif calculated_year >= chk_limit:
                        action_val = 'Repairable' if repairable_val == 'Yes' else 'Non-Repairable'
                
                # Format calculated year
                year_display = str(calculated_year) if calculated_year else ''
                
                # Insert virtual values next to dummy serial
                virtual_values = [
                    actual_barcode,
                    barcode_length,
                    year_display,
                    action_val
                ]
                
                for i, val in enumerate(virtual_values):
                    row.insert(insert_pos + i, val)
                new_rows.append(row)
                
            # Create DataFrame and write to sheet
            df = pd.DataFrame(new_rows[1:], columns=new_rows[0])
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            
        # Rule 2: Append Export History after the original sheets
        if export_history:
            export_hist_cols = [
                "Export Number", 
                "Timestamp", 
                "PCBs Scanned",
                "Who Exported"
            ]
            export_hist_rows = []
            for hist in export_history:
                export_hist_rows.append([
                    hist.get('export_number', 1),
                    hist.get('timestamp', ''),
                    hist.get('scanned_count', 0),
                    hist.get('exported_by', 'Unknown')
                ])
            df_hist = pd.DataFrame(export_hist_rows, columns=export_hist_cols)
            df_hist.to_excel(writer, sheet_name="Export History", index=False)

    # Style sheet look using openpyxl
    wb = openpyxl.load_workbook(dest_path)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    virtual_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    virtual_font = Font(name="Arial", size=10, bold=True, color="1F4E78")
    normal_font = Font(name="Arial", size=10)
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Enable grid lines explicitly
        ws.views.sheetView[0].showGridLines = True
            
        # Standard sheet headers formatting
        is_orig_sheet = (sheet_name not in ["Export History"])
        
        # Format Headers
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_idx)
            val = str(cell.value or '')
            
            # Format virtual headers differently
            if is_orig_sheet and val in ["Actual Serial No", "Length of Actual Serial No", "Mfg Year", "Action"]:
                cell.fill = virtual_fill
                cell.font = virtual_font
            else:
                cell.fill = header_fill
                cell.font = header_font
                
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        # Format Data cells
        for row_idx in range(2, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = normal_font
                
                # Check alignment
                val_str = str(cell.value or '')
                if val_str.isdigit():
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="left", vertical="center")

        # Auto-adjust columns width
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val = str(cell.value or '')
                if val:
                    max_len = max(max_len, len(val))
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(dest_path)

def String_Match(val1, val2):
    return str(val1).lower().strip() == str(val2).lower().strip()

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python export_excel.py <path_to_json_input>")
        sys.exit(1)
        
    with open(sys.argv[1], 'r', encoding='utf-8') as f:
        data = json.load(f)
        
    generate_excel(data)
    print("SUCCESS")
